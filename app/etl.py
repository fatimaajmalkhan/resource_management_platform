"""Reads the Excel file, cleans it, runs quality checks, loads it in,
and (on later runs) logs every real change to History."""
import os
import threading
import time
import numpy as np
import openpyxl
import pandas as pd
from datetime import datetime, date

from app.db import get_session, engine
from app.models import Base, Resource, Deal
from app.quality_rules import run_quality_checks
from app.history import log_change
from app.funnel_history import log_deal_created, log_deal_updated

COLUMN_MAP = {
    "Emp ID": "emp_id", "Resource Name": "resource_name", "Job Title": "job_title",
    "Line Manager": "line_manager", "Line Manager ID": "line_manager_id",
    "Practice": "practice", "Sub-Practice": "sub_practice", "Grade": "grade",
    "Type": "employee_type", "Project / Client / Squad": "project_client_squad",
    "Billable Flag": "billable_flag", "Billable %": "billable_pct",
    "Daily Rate (USD)": "daily_rate_usd", "Days Billed": "days_billed",
    "Monthly Billing (USD)": "monthly_billing_usd", "Engagement Start": "engagement_start",
    "Release Date": "release_date", "Resource Status": "resource_status",
    "Hire Date": "hire_date", "HRBP": "hrbp", "Department": "department",
    "Location Name": "location_name", "Email Address": "email_address",
    "Comments": "comments",
}

# All Excel-mapped fields sync both ways, except emp_id (the join key itself).
TRACKED_FIELDS = [f for f in COLUMN_MAP.values() if f != "emp_id"]

DATE_FIELDS = ["engagement_start", "release_date", "hire_date"]

DB_FIELD_TO_EXCEL_HEADER = {v: k for k, v in COLUMN_MAP.items()}

# Known typo/abbreviation/case variants that all mean the same employee_type category.
EMPLOYEE_TYPE_SYNONYMS = {
    "contract": "Contractual", "contractual": "Contractual",
    "probation": "Probationary", "probationary": "Probationary",
}

def normalize_employee_type(value):
    if value is None:
        return None
    trimmed = str(value).strip()
    return EMPLOYEE_TYPE_SYNONYMS.get(trimmed.lower(), trimmed)

DEFAULT_EXCEL_PATH = os.getenv("EXCEL_PATH", os.path.join(os.path.dirname(__file__), "..", "data",
                                   "Teknosys_Resource_Management_Tool_SAMPLE.xlsx"))

# Allows disabling Excel sync completely in containerized or cloud environments
DISABLE_EXCEL_SYNC = os.getenv("DISABLE_EXCEL_SYNC", "false").lower() == "true"

# Guards all Excel file access (reads and writes) so the background poller
# and platform-triggered writes can't race on the same file.
_excel_file_lock = threading.Lock()


def _retry_excel_write(write_fn, attempts=3, delay_seconds=0.5):
    """Retries write_fn a few times with a short delay -- covers the common
    transient-lock case (a brief external process holding the file, e.g. an
    antivirus/backup scan) so an inline write succeeds within the same
    request instead of waiting for the next poll-loop self-heal. Re-raises
    the last exception if every attempt fails, so callers' existing
    best-effort try/except still reports the failure (e.g. excel_synced)."""
    last_exc = None
    for attempt in range(1, attempts + 1):
        try:
            write_fn()
            return
        except Exception as e:
            last_exc = e
            if attempt < attempts:
                time.sleep(delay_seconds)
    raise last_exc

def _read_and_clean(path):
    raw = pd.read_excel(path, sheet_name="Resource Master")
    df = raw.rename(columns=COLUMN_MAP)
    df = df[[c for c in COLUMN_MAP.values() if c in df.columns]]
    df["billable_flag"] = df["billable_flag"].apply(
        lambda v: None if pd.isna(v) else str(v).strip().lower() in ("yes", "true", "1"))
    if "line_manager_id" in df.columns:
        df["line_manager_id"] = df["line_manager_id"].apply(lambda v: None if pd.isna(v) else int(v))
    if "employee_type" in df.columns:
        df["employee_type"] = df["employee_type"].apply(
            lambda v: None if pd.isna(v) else normalize_employee_type(v))
    for field in DATE_FIELDS:
        if field in df.columns:
            df[field] = df[field].apply(
                lambda v: v.date() if isinstance(v, (pd.Timestamp, datetime)) else (None if pd.isna(v) else v))
    return df.replace({np.nan: None}).to_dict(orient="records")

def check_duplicate_emp_ids(rows):
    """Check and display duplicate emp_ids with their resource names."""
    emp_id_map = {}
    duplicates = {}
    
    for row in rows:
        emp_id = row.get("emp_id")
        resource_name = row.get("resource_name", "N/A")
        
        if emp_id is None:
            continue
            
        if emp_id in emp_id_map:
            if emp_id not in duplicates:
                duplicates[emp_id] = [emp_id_map[emp_id]]
            duplicates[emp_id].append(resource_name)
        else:
            emp_id_map[emp_id] = resource_name
    
    if duplicates:
        print("\nWARNING: DUPLICATE EMP_IDs FOUND:\n")
        for emp_id, names in sorted(duplicates.items()):
            print(f"  Emp ID {emp_id}:")
            for i, name in enumerate(names, 1):
                print(f"    {i}. {name}")
        print(f"\n  Total duplicates: {len(duplicates)}\n")
        return duplicates
    else:
        print("OK: No duplicate emp_ids found.\n")
        return {}

def load_from_excel(path):
    """Run this once, on an empty database."""
    Base.metadata.create_all(engine)
    
    with get_session() as session:
        from sqlalchemy import inspect
        inspector = inspect(engine)
        if inspector.has_table("resources"):
            existing_count = session.query(Resource).count()
            if existing_count > 0:
                print(f"ERROR: The database already contains {existing_count} resources.")
                print("To perform a clean initial load, delete the 'resource_platform.db' database file first.")
                print("If you want to update/sync changes from the Excel file, run this command instead:")
                print("  python scripts/run_sync.py\n")
                return

    rows = _read_and_clean(path)
    
    # Check for duplicates before loading
    check_duplicate_emp_ids(rows)
    
    with get_session() as session:
        for row in rows:
            row["data_flag"] = run_quality_checks(row)
            row["loaded_at"] = datetime.utcnow()
            session.add(Resource(**{k: v for k, v in row.items() if hasattr(Resource, k)}))
            log_change(session, row["emp_id"], "record", None, "created", "initial_load")
        session.commit()
    print(f"Loaded {len(rows)} resources.")

def sync_from_excel(path, source_label=None):
    """Run this every time after -- inserts new people, logs changes to existing ones."""
    if DISABLE_EXCEL_SYNC:
        print("Excel sync is disabled. Skipping sync_from_excel.")
        return
    source_label = source_label or f"sync_{datetime.utcnow().date()}"
    with _excel_file_lock:
        rows = _read_and_clean(path)

    # Check for duplicates before syncing
    check_duplicate_emp_ids(rows)
    
    inserted, updated_fields = 0, 0
    with get_session() as session:
        for row in rows:
            row["data_flag"] = run_quality_checks(row)
            existing = session.get(Resource, row["emp_id"])
            if existing is None:
                row["loaded_at"] = datetime.utcnow()
                session.add(Resource(**{k: v for k, v in row.items() if hasattr(Resource, k)}))
                log_change(session, row["emp_id"], "record", None, "created", source_label)
                inserted += 1
                continue
            for field in TRACKED_FIELDS:
                old_value, new_value = getattr(existing, field), row.get(field)
                if old_value != new_value:
                    log_change(session, row["emp_id"], field, old_value, new_value, source_label)
                    setattr(existing, field, new_value)
                    updated_fields += 1
            existing.data_flag = row["data_flag"]
        session.commit()
    print(f"Sync complete. {inserted} new, {updated_fields} field change(s) logged.")

def append_resource_to_excel(record: dict, path: str = None):
    """Append one newly-created resource as a row in the source Excel file,
    so it stays current with resources added through the app. One-way
    (DB -> Excel), append-only. Retries a few times on transient failures
    (e.g. the file briefly locked by another process) before giving up --
    callers should still treat a final failure as non-fatal (the 20s poll
    loop's reconcile_missing_resources_to_excel will catch it later)."""
    if DISABLE_EXCEL_SYNC:
        return
    path = path or DEFAULT_EXCEL_PATH

    header_row_values = {
        DB_FIELD_TO_EXCEL_HEADER.get(field): value
        for field, value in record.items()
        if field in DB_FIELD_TO_EXCEL_HEADER
    }
    if "Billable Flag" in header_row_values:
        flag = header_row_values["Billable Flag"]
        header_row_values["Billable Flag"] = None if flag is None else ("Yes" if flag else "No")
    # Columns that exist in the sheet but aren't in COLUMN_MAP -- mirror the
    # convention already used by scripts/generate_6month_data.py.
    header_row_values["Level"] = record.get("grade")
    header_row_values["Employee Category"] = record.get("employee_type")

    def _write():
        with _excel_file_lock:
            wb = openpyxl.load_workbook(path)
            ws = wb["Resource Master"]
            header_row = [cell.value for cell in ws[1]]
            row_values = [header_row_values.get(col) for col in header_row]
            ws.append(row_values)
            wb.save(path)

    _retry_excel_write(_write)

def update_resource_in_excel(record: dict, path: str = None):
    """Update an existing resource's row in the source Excel file in place,
    so edits made in the platform (e.g. via the chatbot's apply_change) stay
    reflected in the spreadsheet. One-way (DB -> Excel), best-effort --
    callers should treat failures (e.g. the file being open elsewhere) as
    non-fatal. Raises if no row matches the given emp_id."""
    if DISABLE_EXCEL_SYNC:
        return
    path = path or DEFAULT_EXCEL_PATH

    header_row_values = {
        DB_FIELD_TO_EXCEL_HEADER.get(field): value
        for field, value in record.items()
        if field in DB_FIELD_TO_EXCEL_HEADER
    }
    if "Billable Flag" in header_row_values:
        flag = header_row_values["Billable Flag"]
        header_row_values["Billable Flag"] = None if flag is None else ("Yes" if flag else "No")
    header_row_values["Level"] = record.get("grade")
    header_row_values["Employee Category"] = record.get("employee_type")

    with _excel_file_lock:
        wb = openpyxl.load_workbook(path)
        ws = wb["Resource Master"]
        header_row = [cell.value for cell in ws[1]]
        emp_id_col = header_row.index("Emp ID") + 1  # openpyxl columns are 1-indexed

        target_row_idx = None
        for row in ws.iter_rows(min_row=2):
            if row[emp_id_col - 1].value == record.get("emp_id"):
                target_row_idx = row[0].row
                break
        if target_row_idx is None:
            raise ValueError(f"No Excel row found for emp_id {record.get('emp_id')}")

        for col_idx, header in enumerate(header_row, start=1):
            if header in header_row_values:
                ws.cell(row=target_row_idx, column=col_idx, value=header_row_values[header])
        wb.save(path)

def remove_resource_from_excel(emp_id: int, path: str = None):
    """Remove a resource's row from the source Excel file when it's deleted
    in the platform. One-way (DB -> Excel), best-effort, retried on transient
    failures. Without this, a deleted resource's row would linger in Excel
    and get misread as "new" by the next sync_from_excel poll, silently
    recreating it in the DB -- exactly the bug this closes. A no-op (not an
    error) if no matching row is found, e.g. it was already removed."""
    if DISABLE_EXCEL_SYNC:
        return
    path = path or DEFAULT_EXCEL_PATH

    def _write():
        with _excel_file_lock:
            wb = openpyxl.load_workbook(path)
            ws = wb["Resource Master"]
            header_row = [cell.value for cell in ws[1]]
            emp_id_col = header_row.index("Emp ID") + 1

            target_row_idx = None
            for row in ws.iter_rows(min_row=2):
                if row[emp_id_col - 1].value == emp_id:
                    target_row_idx = row[0].row
                    break
            if target_row_idx is None:
                return
            ws.delete_rows(target_row_idx)
            wb.save(path)

    _retry_excel_write(_write)

def reconcile_missing_resources_to_excel(path: str = None) -> int:
    """Self-heal: find DB resources with no matching Excel row (e.g. from a
    transient append failure at creation time, like the file being open in
    Excel) and append them. Missing-row detection only -- does not diff
    field-level drift for rows that already exist in both places."""
    if DISABLE_EXCEL_SYNC:
        return 0
    path = path or DEFAULT_EXCEL_PATH

    with get_session() as session:
        resources_by_id = {r.emp_id: r for r in session.query(Resource).all()}

    with _excel_file_lock:
        wb = openpyxl.load_workbook(path)
        ws = wb["Resource Master"]
        header_row = [cell.value for cell in ws[1]]
        emp_id_col = header_row.index("Emp ID")
        excel_emp_ids = {
            row[emp_id_col].value for row in ws.iter_rows(min_row=2)
            if row[emp_id_col].value is not None
        }

    missing_ids = set(resources_by_id) - excel_emp_ids
    fixed = 0
    for emp_id in missing_ids:
        r = resources_by_id[emp_id]
        record = {"emp_id": emp_id, **{f: getattr(r, f) for f in DB_FIELD_TO_EXCEL_HEADER}}
        try:
            append_resource_to_excel(record, path=path)
            fixed += 1
        except Exception as e:
            print(f"Reconcile: failed to append missing emp_id {emp_id} to Excel: {e}")
    return fixed


# ---------------------------------------------------------------------------
# Funnel sheet <-> Deal table sync. The Funnel sheet is a wide/pivoted format
# (one row can carry demand for up to 4 roles via separate quantity columns),
# unlike Resource Master's one-row-per-record shape, so it needs its own
# read/write logic rather than reusing _read_and_clean.
# ---------------------------------------------------------------------------
FUNNEL_SHEET = "Funnel"
FUNNEL_HEADER_ROW = 3  # 0-indexed row for pandas' header= (row 4 in Excel)

FUNNEL_COLUMN_MAP = {
    "#": "seq",
    "Client / Project": "client_project",
    "Stage": "stage",
    "Probability": "probability",
    "Pool / Type": "practice",
    "Eng. Start\n(Est.)": "eng_start_est",
    "BI": "BI",
    "DBA": "DBA",
    "Data\nEngineer": "Data Engineer",
    "Other": "Other",
    "Total\nResources": "total_resources",
    "Target\nMonth": "target_month",
    "Notes": "notes",
}

# The 4 wide-format role columns, after renaming -- these match Deal.role
# values exactly (and what match_role() in server.py expects).
ROLE_FIELDS = ["BI", "DBA", "Data Engineer", "Other"]

# Clean role name -> raw Excel header (inverse of the relevant part of
# FUNNEL_COLUMN_MAP), for locating/writing a specific role's column.
ROLE_TO_HEADER = {"BI": "BI", "DBA": "DBA", "Data Engineer": "Data\nEngineer", "Other": "Other"}

# Fields tracked for diffing on an existing (client_project, role, target_month)
# match. client_project/role/target_month are the identity key, not tracked.
TRACKED_DEAL_FIELDS = ["stage", "probability", "quantity", "practice", "notes"]


def normalize_probability(v):
    """The Funnel sheet stores Probability as a 0-1 fraction (0.7, 1.0, ...)
    but Deal.probability is 0-100. Tolerates a future whole-number entry
    (e.g. 70) too -- anything already > 1 is assumed to already be a
    percentage."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    v = float(v)
    return v * 100 if 0 <= v <= 1 else v


def normalize_target_month(v):
    """Coerces a Target Month cell to a 'YYYY-MM' string, or None if it
    can't be parsed (a date object, or a plain 'YYYY-MM' string are both
    accepted)."""
    if v is None or (isinstance(v, float) and pd.isna(v)):
        return None
    if isinstance(v, (pd.Timestamp, datetime, date)):
        return v.strftime("%Y-%m")
    s = str(v).strip()
    try:
        datetime.strptime(s, "%Y-%m")
        return s
    except ValueError:
        return None


def _read_funnel_rows(path):
    """Reads the Funnel sheet and decomposes each wide row into flat
    per-role deal dicts (0-4 per row, one per positive-quantity role column).

    Stops at the first row whose '#' isn't a positive integer -- verified
    firsthand that the sheet has a blank separator row immediately after the
    real deals, followed by a 'TOTAL PIPELINE DEMAND' summary block and a
    'STATUS LEGEND' block further down (formatting extends to row 100). That
    first blank-'#' row is the only reliable end-of-table sentinel; treating
    a later non-numeric '#' as "skip and keep scanning" would eventually
    misparse the summary/legend rows as deals."""
    raw = pd.read_excel(path, sheet_name=FUNNEL_SHEET, header=FUNNEL_HEADER_ROW)
    df = raw.rename(columns=FUNNEL_COLUMN_MAP)

    rows = []
    for _, row in df.iterrows():
        seq = row.get("seq")
        try:
            seq_ok = seq is not None and not pd.isna(seq) and float(seq) > 0
        except (TypeError, ValueError):
            seq_ok = False
        if not seq_ok:
            break

        raw_client = row.get("client_project")
        client_project = None if pd.isna(raw_client) else str(raw_client).strip()
        if not client_project:
            print(f"Funnel sync: skipping row #{seq} -- blank Client / Project")
            continue

        target_month = normalize_target_month(row.get("target_month"))
        if target_month is None:
            print(f"Funnel sync: skipping row #{seq} ({client_project}) -- unparseable Target Month")
            continue

        raw_stage = row.get("stage")
        stage = None if pd.isna(raw_stage) else str(raw_stage).strip()
        probability = normalize_probability(row.get("probability"))
        raw_practice = row.get("practice")
        practice = None if pd.isna(raw_practice) else str(raw_practice).strip()
        raw_notes = row.get("notes")
        notes = None if pd.isna(raw_notes) else str(raw_notes).strip()

        for role in ROLE_FIELDS:
            qty_raw = row.get(role)
            if qty_raw is None or pd.isna(qty_raw):
                continue
            qty = int(qty_raw)
            if qty <= 0:
                continue
            rows.append({
                "client_project": client_project, "role": role, "target_month": target_month,
                "stage": stage, "probability": probability, "practice": practice,
                "quantity": qty, "notes": notes,
            })
    return rows


def sync_deals_from_excel(path=None, source_label=None):
    """Mirrors sync_from_excel, scoped to Deal: upserts rows from the Funnel
    sheet keyed on (client_project, role, target_month) -- there's no stable
    ID column in the sheet, so this composite is the identity. Inserts new
    combinations (logged via log_deal_created) and diffs+logs each changed
    field on existing ones (log_deal_updated); never deletes a Deal whose
    row disappears or gets zeroed out in the sheet, matching the additive
    philosophy sync_from_excel already uses for resources."""
    if DISABLE_EXCEL_SYNC:
        print("Excel sync is disabled. Skipping sync_deals_from_excel.")
        return
    path = path or DEFAULT_EXCEL_PATH
    source_label = source_label or f"funnel_sync_{datetime.utcnow().date()}"

    with _excel_file_lock:
        rows = _read_funnel_rows(path)

    inserted, updated_fields = 0, 0
    with get_session() as session:
        for row in rows:
            existing = (
                session.query(Deal)
                .filter_by(client_project=row["client_project"], role=row["role"],
                           target_month=row["target_month"])
                .first()
            )
            if existing is None:
                new_deal = Deal(
                    client_project=row["client_project"], stage=row["stage"],
                    probability=row["probability"], role=row["role"],
                    quantity=row["quantity"], target_month=row["target_month"],
                    practice=row["practice"], notes=row["notes"],
                )
                session.add(new_deal)
                session.commit()
                session.refresh(new_deal)
                log_deal_created(session, new_deal, source=source_label)
                session.commit()
                inserted += 1
                continue

            new_values = {"stage": row["stage"], "probability": row["probability"],
                          "quantity": row["quantity"], "practice": row["practice"],
                          "notes": row["notes"]}
            for field, new_value in new_values.items():
                old_value = getattr(existing, field)
                if old_value != new_value:
                    log_deal_updated(session, existing, field, old_value, new_value, source=source_label)
                    setattr(existing, field, new_value)
                    updated_fields += 1
            session.commit()

    print(f"Funnel sync complete. {inserted} new deal(s), {updated_fields} field change(s) logged.")


def append_deal_to_excel(record: dict, path: str = None):
    """Append one newly-created deal as a row in the Funnel sheet. One-way
    (DB -> Excel), mirroring append_resource_to_excel -- including the retry
    on transient failures -- but can't reuse ws.append(): the Funnel sheet's
    formatting extends to row 100 while real data ends around row 10 (a
    blank separator, a 'TOTAL PIPELINE DEMAND' summary, and a 'STATUS LEGEND'
    block all sit below it), so appending would land the new row far below
    the visible table. Instead this finds the real last data row via the
    same '#' sentinel _read_funnel_rows uses, and inserts a row right after
    it. A final failure (after retries) is non-fatal -- the 20s poll loop's
    reconcile_missing_deals_to_excel will catch it later."""
    if DISABLE_EXCEL_SYNC:
        return
    path = path or DEFAULT_EXCEL_PATH

    def _write():
        with _excel_file_lock:
            wb = openpyxl.load_workbook(path)
            ws = wb[FUNNEL_SHEET]
            header_row_idx = FUNNEL_HEADER_ROW + 1  # openpyxl rows are 1-indexed
            header_row = [cell.value for cell in ws[header_row_idx]]
            seq_col = header_row.index("#") + 1

            last_data_row = header_row_idx
            max_seq = 0
            for row in ws.iter_rows(min_row=header_row_idx + 1):
                seq = row[seq_col - 1].value
                try:
                    seq_ok = seq is not None and float(seq) > 0
                except (TypeError, ValueError):
                    seq_ok = False
                if not seq_ok:
                    break
                last_data_row = row[0].row
                max_seq = max(max_seq, int(seq))

            new_row_idx = last_data_row + 1
            ws.insert_rows(new_row_idx)

            role = record.get("role")
            role_quantities = {r: 0 for r in ROLE_FIELDS}
            if role in role_quantities:
                role_quantities[role] = record.get("quantity")

            header_to_value = {
                "#": max_seq + 1,
                "Client / Project": record.get("client_project"),
                "Stage": record.get("stage"),
                "Probability": (record.get("probability") or 0) / 100.0,
                "Pool / Type": record.get("practice"),
                "Eng. Start\n(Est.)": None,
                "BI": role_quantities["BI"],
                "DBA": role_quantities["DBA"],
                "Data\nEngineer": role_quantities["Data Engineer"],
                "Other": role_quantities["Other"],
                "Total\nResources": None,
                "Target\nMonth": record.get("target_month"),
                "Notes": record.get("notes"),
            }

            for col_idx, header in enumerate(header_row, start=1):
                if header in header_to_value:
                    ws.cell(row=new_row_idx, column=col_idx, value=header_to_value[header])

            wb.save(path)

    _retry_excel_write(_write)


def remove_deal_from_excel(client_project: str, role: str, target_month: str, path: str = None):
    """Remove a deal's contribution from the Funnel sheet when it's deleted
    in the platform. One-way (DB -> Excel), retried on transient failures.
    Without this, a deleted deal's row would linger in Excel and get
    misread as "new" by the next sync_deals_from_excel poll, silently
    recreating it in the DB. Since one Funnel row can carry demand for up to
    4 roles, this only zeroes out the matching role's quantity column, and
    deletes the whole row only if all 4 role columns are then zero -- so
    deleting one role's deal doesn't erase a sibling role's still-active
    demand on the same row. A no-op if no matching row is found."""
    if DISABLE_EXCEL_SYNC:
        return
    path = path or DEFAULT_EXCEL_PATH

    def _write():
        with _excel_file_lock:
            wb = openpyxl.load_workbook(path)
            ws = wb[FUNNEL_SHEET]
            header_row_idx = FUNNEL_HEADER_ROW + 1
            header_row = [cell.value for cell in ws[header_row_idx]]
            seq_col = header_row.index("#") + 1
            proj_col = header_row.index("Client / Project") + 1
            month_col = header_row.index("Target\nMonth") + 1
            role_col = header_row.index(ROLE_TO_HEADER[role]) + 1

            target_row_idx = None
            for row in ws.iter_rows(min_row=header_row_idx + 1):
                seq = row[seq_col - 1].value
                try:
                    seq_ok = seq is not None and float(seq) > 0
                except (TypeError, ValueError):
                    seq_ok = False
                if not seq_ok:
                    break
                proj_val = row[proj_col - 1].value
                month_val = row[month_col - 1].value
                if (proj_val is not None and str(proj_val).strip() == client_project
                        and str(month_val).strip() == target_month):
                    target_row_idx = row[0].row
                    break

            if target_row_idx is None:
                return

            ws.cell(row=target_row_idx, column=role_col, value=0)

            all_role_cols = [header_row.index(h) + 1 for h in ROLE_TO_HEADER.values()]
            remaining = [ws.cell(row=target_row_idx, column=c).value or 0 for c in all_role_cols]
            if sum(remaining) <= 0:
                ws.delete_rows(target_row_idx)

            wb.save(path)

    _retry_excel_write(_write)


def reconcile_missing_deals_to_excel(path: str = None) -> int:
    """Self-heal: find DB deals with no matching Funnel-sheet row (e.g. from
    a transient append failure at creation time) and append them. Mirrors
    reconcile_missing_resources_to_excel, but keyed on the same
    (client_project, role, target_month) composite sync_deals_from_excel
    uses, since the sheet has no stable ID column. Missing-row detection
    only -- does not diff field-level drift for deals that already exist in
    both places (that's sync_deals_from_excel's job)."""
    if DISABLE_EXCEL_SYNC:
        return 0
    path = path or DEFAULT_EXCEL_PATH

    with get_session() as session:
        deals = session.query(Deal).all()

    with _excel_file_lock:
        existing_rows = _read_funnel_rows(path)
    existing_keys = {(r["client_project"], r["role"], r["target_month"]) for r in existing_rows}

    fixed = 0
    for d in deals:
        key = (d.client_project, d.role, d.target_month)
        if key in existing_keys:
            continue
        record = {
            "client_project": d.client_project, "stage": d.stage,
            "probability": d.probability, "role": d.role,
            "quantity": d.quantity, "target_month": d.target_month,
            "practice": d.practice, "notes": d.notes,
        }
        try:
            append_deal_to_excel(record, path=path)
            fixed += 1
        except Exception as e:
            print(f"Reconcile: failed to append missing deal '{d.client_project}' "
                  f"({d.role}, {d.target_month}) to Excel: {e}")
    return fixed