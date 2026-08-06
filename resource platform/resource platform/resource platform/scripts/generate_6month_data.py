"""Add 42 new employee rows (Jan–Jun 2026) to the Excel Resource Master sheet.

Run:  python scripts/generate_6month_data.py
Then: python scripts/run_sync.py
"""
import sys, os
sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

import openpyxl
from datetime import date, timedelta

EXCEL_PATH = os.path.join(os.path.dirname(__file__), "..", "data",
                          "Teknosys_Resource_Management_Tool_SAMPLE.xlsx")
SHEET_NAME = "Resource Master"

# ---------------------------------------------------------------------------
# Rate lookup
# ---------------------------------------------------------------------------
GRADE_RATE = {"L1": 150, "L2": 200, "L3": 250, "L4": 350, "L5": 500}
CONTRACTUAL_RATE = 180

def daily_rate(grade, emp_type):
    if "Contractual" in emp_type:
        return CONTRACTUAL_RATE
    return GRADE_RATE.get(grade, 200)

def monthly_billing(rate, billable):
    return rate * 20 if billable else 0

def engagement_start(hire_date):
    return hire_date + timedelta(days=14)

def email(name):
    parts = name.lower().split()
    return f"{parts[0]}.{parts[-1]}@teknosys.com"

# ---------------------------------------------------------------------------
# 42 new employees  (name, job_title, practice, sub_practice, grade, emp_type,
#                    project, location, hire_date, release_date, resource_status)
# ---------------------------------------------------------------------------
_A = "Analytics & Insights"
_S = "Software - TSF"
_FT = "Full-Time Regular"
_CT = "Contractual"
_PR = "Probationary"

HRBP_POOL = ["Amna Zahid", "Khalid Mehmood"]
MGR_POOL  = [("Zaid Hassan", 1001), ("Sara Khan", 1002),
             ("Ahmed Raza", 1003), ("Nadia Ahmed", 1004)]

NEW_EMPLOYEES = [
    # ── January 2026 ──────────────────────────────────────────────────────
    ("Ahsan Ahmed",     "Data Engineer",       _A, "Data Engineering",       "L3", _FT, "ClientCo - Alpha",  "Karachi",   date(2026,1,3),  None,            "Active"),
    ("Bilal Khan",      "BI Analyst",          _A, "Business Intelligence",  "L2", _FT, "ClientCo - Beta",   "Lahore",    date(2026,1,7),  None,            "Active"),
    ("Daniyal Malik",   "Software Engineer",   _S, "Backend Development",    "L2", _FT, "ClientCo - Gamma",  "Karachi",   date(2026,1,10), None,            "Active"),
    ("Faisal Sheikh",   "Analytics Engineer",  _A, "Analytics Engineering",  "L3", _CT, "ClientCo - Alpha",  "Dubai",     date(2026,1,14), date(2026,5,15), "Released"),
    ("Hassan Qureshi",  "QA Engineer",         _S, "QA Engineering",         "L1", _FT, "ClientCo - Gamma",  "Islamabad", date(2026,1,17), None,            "Active"),
    ("Imran Siddiqui",  "Data Engineer",       _A, "Data Engineering",       "L4", _FT, "ClientCo - Beta",   "Karachi",   date(2026,1,21), None,            "Active"),
    ("Junaid Mirza",    "Software Engineer",   _S, "Frontend Development",   "L2", _FT, "",                  "Lahore",    date(2026,1,28), None,            "Active"),

    # ── February 2026 ─────────────────────────────────────────────────────
    ("Kamran Butt",     "DBA",                 _A, "Database",               "L3", _FT, "ClientCo - Alpha",  "Karachi",   date(2026,2,2),  None,            "Active"),
    ("Luqman Chaudhry", "BI Analyst",          _A, "Business Intelligence",  "L2", _PR, "ClientCo - Beta",   "Islamabad", date(2026,2,9),  None,            "Active"),
    ("Mohsin Rashid",   "Data Engineer",       _A, "Data Engineering",       "L1", _FT, "ClientCo - Alpha",  "Lahore",    date(2026,2,13), date(2026,4,30), "Released"),
    ("Nawaz Hussain",   "Software Engineer",   _S, "Backend Development",    "L3", _FT, "ClientCo - Gamma",  "Karachi",   date(2026,2,17), None,            "Active"),
    ("Omar Akhtar",     "Analytics Engineer",  _A, "Analytics Engineering",  "L2", _CT, "",                  "Dubai",     date(2026,2,21), None,            "Active"),
    ("Qasim Baig",      "Software Engineer",   _S, "Frontend Development",   "L4", _FT, "ClientCo - Gamma",  "Karachi",   date(2026,2,25), None,            "Active"),

    # ── March 2026 ────────────────────────────────────────────────────────
    ("Raza Cheema",     "Data Engineer",       _A, "Data Engineering",       "L3", _FT, "ClientCo - Alpha",  "Lahore",    date(2026,3,2),  None,            "Active"),
    ("Saad Dogar",      "BI Analyst",          _A, "Business Intelligence",  "L5", _FT, "ClientCo - Beta",   "Karachi",   date(2026,3,5),  None,            "Active"),
    ("Tariq Farooqi",   "Software Engineer",   _S, "Backend Development",    "L2", _FT, "ClientCo - Gamma",  "Islamabad", date(2026,3,9),  None,            "Active"),
    ("Usman Gillani",   "DBA",                 _A, "Database",               "L3", _CT, "ClientCo - Alpha",  "Dubai",     date(2026,3,12), None,            "Active"),
    ("Waqar Hashmi",    "Data Engineer",       _A, "Data Engineering",       "L2", _FT, "ClientCo - Beta",   "Karachi",   date(2026,3,16), None,            "Active"),
    ("Yasir Ijaz",      "QA Engineer",         _S, "QA Engineering",         "L1", _FT, "",                  "Lahore",    date(2026,3,19), date(2026,6,1),  "Released"),
    ("Zain Jaffery",    "Analytics Engineer",  _A, "Analytics Engineering",  "L3", _FT, "ClientCo - Alpha",  "Karachi",   date(2026,3,23), None,            "Active"),
    ("Asad Kazmi",      "Software Engineer",   _S, "Frontend Development",   "L2", _PR, "ClientCo - Gamma",  "Islamabad", date(2026,3,28), None,            "Active"),

    # ── April 2026 ────────────────────────────────────────────────────────
    ("Babar Lodhi",     "Data Engineer",       _A, "Data Engineering",       "L4", _FT, "ClientCo - Beta",   "Karachi",   date(2026,4,1),  None,            "Active"),
    ("Dawood Minhas",   "BI Analyst",          _A, "Business Intelligence",  "L3", _FT, "ClientCo - Alpha",  "Lahore",    date(2026,4,6),  None,            "Active"),
    ("Farhan Niazi",    "Software Engineer",   _S, "DevOps",                 "L3", _FT, "ClientCo - Gamma",  "Dubai",     date(2026,4,10), None,            "Active"),
    ("Haris Naqvi",     "Analytics Engineer",  _A, "Analytics Engineering",  "L2", _CT, "",                  "Karachi",   date(2026,4,14), None,            "Active"),
    ("Irfan Omer",      "Data Engineer",       _A, "Data Engineering",       "L1", _FT, "ClientCo - Beta",   "Islamabad", date(2026,4,18), None,            "Active"),
    ("Jawad Pasha",     "DBA",                 _A, "Database",               "L3", _FT, "ClientCo - Alpha",  "Karachi",   date(2026,4,22), None,            "Active"),
    ("Kaleem Qazi",     "Software Engineer",   _S, "Backend Development",    "L2", _FT, "ClientCo - Gamma",  "Lahore",    date(2026,4,25), None,            "Active"),

    # ── May 2026 ──────────────────────────────────────────────────────────
    ("Mubarak Raja",    "Data Engineer",       _A, "Data Engineering",       "L3", _FT, "ClientCo - Beta",   "Karachi",   date(2026,5,4),  None,            "Active"),
    ("Naeem Sardar",    "BI Analyst",          _A, "Business Intelligence",  "L2", _FT, "ClientCo - Alpha",  "Islamabad", date(2026,5,7),  None,            "Active"),
    ("Owais Toor",      "Software Engineer",   _S, "Frontend Development",   "L3", _CT, "ClientCo - Gamma",  "Dubai",     date(2026,5,12), None,            "Active"),
    ("Parvez Ullah",    "Analytics Engineer",  _A, "Analytics Engineering",  "L4", _FT, "ClientCo - Beta",   "Karachi",   date(2026,5,16), None,            "Active"),
    ("Rafay Vaqar",     "Data Engineer",       _A, "Data Engineering",       "L2", _FT, "",                  "Lahore",    date(2026,5,19), None,            "Active"),
    ("Salman Wasim",    "QA Engineer",         _S, "QA Engineering",         "L2", _FT, "ClientCo - Gamma",  "Karachi",   date(2026,5,23), None,            "Active"),
    ("Talha Yahya",     "Software Engineer",   _S, "Backend Development",    "L1", _FT, "ClientCo - Gamma",  "Islamabad", date(2026,5,29), None,            "Active"),

    # ── June 2026 ─────────────────────────────────────────────────────────
    ("Umer Zafar",      "Data Engineer",       _A, "Data Engineering",       "L3", _FT, "ClientCo - Alpha",  "Karachi",   date(2026,6,1),  None,            "Active"),
    ("Waseem Zaidi",    "BI Analyst",          _A, "Business Intelligence",  "L2", _PR, "ClientCo - Beta",   "Lahore",    date(2026,6,5),  None,            "Active"),
    ("Yousaf Anjum",    "DBA",                 _A, "Database",               "L3", _CT, "",                  "Islamabad", date(2026,6,10), None,            "Active"),
    ("Zubair Bhatti",   "Software Engineer",   _S, "Frontend Development",   "L2", _FT, "ClientCo - Gamma",  "Karachi",   date(2026,6,14), None,            "Active"),
    ("Adil Chattha",    "Analytics Engineer",  _A, "Analytics Engineering",  "L4", _FT, "ClientCo - Alpha",  "Dubai",     date(2026,6,18), None,            "Active"),
    ("Basit Dhaliwal",  "Data Engineer",       _A, "Data Engineering",       "L2", _FT, "ClientCo - Beta",   "Karachi",   date(2026,6,23), None,            "Active"),
    ("Ghazal Farooq",   "Software Engineer",   _S, "Backend Development",    "L3", _FT, "ClientCo - Gamma",  "Lahore",    date(2026,6,27), None,            "Active"),
]


def build_record(emp_id, emp_data, mgr, hrbp):
    """Return a dict keyed by Excel column names.
    Only the keys that match ETL's COLUMN_MAP are read; extras default to None.
    """
    (name, job_title, practice, sub_practice, grade, emp_type,
     project, location, hire_dt, release_dt, status) = emp_data

    billable = bool(project)
    rate = daily_rate(grade, emp_type)
    billing = monthly_billing(rate, billable)
    eng_start = engagement_start(hire_dt)
    department = "Analytics" if practice == _A else "Software Engineering"

    return {
        "Emp ID":                   emp_id,
        "Resource Name":            name,
        "Job Title":                job_title,
        "Line Manager":             mgr[0],
        "Line Manager ID":          mgr[1],
        "Practice":                 practice,
        "Sub-Practice":             sub_practice,
        "Grade":                    grade,
        "Type":                     emp_type,
        "Project / Client / Squad": project,
        "Billable Flag":            "Yes" if billable else "No",
        "Billable %":               100 if billable else 0,
        "Daily Rate (USD)":         rate,
        "Days Billed":              20 if billable else 0,
        "Monthly Billing (USD)":    billing,
        "Engagement Start":         eng_start,
        "Release Date":             release_dt,
        "Resource Status":          status,
        "Hire Date":                hire_dt,
        "HRBP":                     hrbp,
        "Department":               department,
        "Location Name":            location,
        "Email Address":            email(name),
        "Comments":                 "",
        # Extra columns present in this Excel (not used by ETL, left blank)
        "Sub project/Squad":        None,
        "Assignment Status":        None,
        "Div Head ID":              None,
        "Divisional Head":          None,
        "HOD ID":                   None,
        "HOD":                      None,
        "HRBP ID":                  None,
        "Employee Category":        emp_type,
        "Division":                 None,
        "Sub Department":           None,
        "Region Name":              None,
        "Phone":                    None,
        "Gender":                   None,
        "Level":                    grade,   # mirror of Grade
    }


def main():
    path = os.path.abspath(EXCEL_PATH)
    if not os.path.exists(path):
        print(f"ERROR: Excel file not found at {path}")
        sys.exit(1)

    wb = openpyxl.load_workbook(path)
    if SHEET_NAME not in wb.sheetnames:
        print(f"ERROR: Sheet '{SHEET_NAME}' not found. Available: {wb.sheetnames}")
        sys.exit(1)

    ws = wb[SHEET_NAME]

    # Read actual header row to know column order
    header_row = [cell.value for cell in ws[1]]
    print(f"Detected columns: {header_row}")

    if "Emp ID" not in header_row:
        print("ERROR: 'Emp ID' column not found in header row.")
        sys.exit(1)

    emp_id_idx = header_row.index("Emp ID")   # 0-based

    # Find max existing Emp ID
    max_id = 0
    for row in ws.iter_rows(min_row=2, values_only=True):
        val = row[emp_id_idx]
        if val and isinstance(val, (int, float)):
            max_id = max(max_id, int(val))

    print(f"Current max Emp ID in sheet: {max_id}")
    start_id = max_id + 1

    added = 0
    for i, emp_data in enumerate(NEW_EMPLOYEES):
        emp_id  = start_id + i
        mgr     = MGR_POOL[i % len(MGR_POOL)]
        hrbp    = HRBP_POOL[i % len(HRBP_POOL)]
        record  = build_record(emp_id, emp_data, mgr, hrbp)

        # Build a row list aligned with the actual header order
        row_values = [record.get(col, None) for col in header_row]
        ws.append(row_values)
        added += 1
        print(f"  Added emp_id={emp_id}: {emp_data[0]} ({emp_data[4]}, {emp_data[8].strftime('%b %Y')})")

    wb.save(path)
    print(f"\nDone — added {added} rows to '{SHEET_NAME}' in {path}")
    print("Next step: run `python scripts/run_sync.py` to load into the database.")


if __name__ == "__main__":
    main()
